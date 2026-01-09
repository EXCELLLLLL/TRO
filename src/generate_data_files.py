"""
Data Files Generator
PT. MediCare Indonesia - Transportation Problem

This script generates all data files needed for the project:
1. input_data.xlsx - Complete dataset in Excel format
2. warehouse_capacity.csv - Supply data
3. destination_demand.csv - Demand data
4. transportation_cost.csv - Cost matrix

Run this script first to create all necessary data files.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class DataGenerator:
    """Generate all data files for the transportation problem"""

    def __init__(self):
        """Initialize with transportation data"""

        # Warehouse data
        self.warehouses = ['Jakarta', 'Tangerang', 'Bekasi', 'Bogor']
        self.supply = {
            'Jakarta': 350,
            'Tangerang': 400,
            'Bekasi': 300,
            'Bogor': 250
        }

        # Destination data
        self.destinations = ['RS_Jakarta_Pusat', 'RS_Tangerang', 'RS_Bekasi',
                             'Apotek_Depok', 'RS_Bogor']
        self.demand = {
            'RS_Jakarta_Pusat': 250,
            'RS_Tangerang': 300,
            'RS_Bekasi': 200,
            'Apotek_Depok': 280,
            'RS_Bogor': 220
        }

        # Transportation costs (Rp thousands per unit)
        self.costs = {
            ('Jakarta', 'RS_Jakarta_Pusat'): 5,
            ('Jakarta', 'RS_Tangerang'): 15,
            ('Jakarta', 'RS_Bekasi'): 12,
            ('Jakarta', 'Apotek_Depok'): 8,
            ('Jakarta', 'RS_Bogor'): 18,

            ('Tangerang', 'RS_Jakarta_Pusat'): 18,
            ('Tangerang', 'RS_Tangerang'): 4,
            ('Tangerang', 'RS_Bekasi'): 20,
            ('Tangerang', 'Apotek_Depok'): 16,
            ('Tangerang', 'RS_Bogor'): 25,

            ('Bekasi', 'RS_Jakarta_Pusat'): 14,
            ('Bekasi', 'RS_Tangerang'): 22,
            ('Bekasi', 'RS_Bekasi'): 6,
            ('Bekasi', 'Apotek_Depok'): 10,
            ('Bekasi', 'RS_Bogor'): 20,

            ('Bogor', 'RS_Jakarta_Pusat'): 16,
            ('Bogor', 'RS_Tangerang'): 24,
            ('Bogor', 'RS_Bekasi'): 18,
            ('Bogor', 'Apotek_Depok'): 7,
            ('Bogor', 'RS_Bogor'): 5
        }

        # Create data directory if not exists
        self.data_dir = '../data'
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
            print(f"✓ Created directory: {self.data_dir}")

    def generate_warehouse_capacity_csv(self):
        """Generate warehouse_capacity.csv"""

        filename = f'{self.data_dir}/warehouse_capacity.csv'

        # Create DataFrame
        data = []
        for warehouse in self.warehouses:
            data.append({
                'Warehouse_ID': f'WH_{warehouse[:3].upper()}',
                'Warehouse_Name': warehouse,
                'Location': warehouse,
                'Capacity_Units': self.supply[warehouse],
                'Current_Stock': self.supply[warehouse],
                'Utilization_Percent': 0,  # Will be filled after optimization
                'Manager': f'{warehouse} Manager',
                'Contact': f'{warehouse.lower()}@medicare.co.id',
                'Operating_Hours': '07:00-17:00',
                'Status': 'Active'
            })

        df = pd.DataFrame(data)
        df.to_csv(filename, index=False)

        print(f"✓ Generated: {filename}")
        print(f"  Rows: {len(df)}")
        print(f"  Columns: {len(df.columns)}")

        return df

    def generate_destination_demand_csv(self):
        """Generate destination_demand.csv"""

        filename = f'{self.data_dir}/destination_demand.csv'

        # Create DataFrame
        data = []
        for i, destination in enumerate(self.destinations, 1):
            # Determine type
            if 'RS' in destination:
                facility_type = 'Hospital'
            else:
                facility_type = 'Pharmacy'

            data.append({
                'Destination_ID': f'DEST_{i:03d}',
                'Destination_Name': destination.replace('_', ' '),
                'Facility_Type': facility_type,
                'Location': destination.replace('_', ' '),
                'Monthly_Demand_Units': self.demand[destination],
                'Priority_Level': 'High' if self.demand[destination] >= 250 else 'Medium',
                'Service_Days': 'Mon-Sat',
                'Delivery_Window': '08:00-15:00',
                'Contact_Person': f'{destination.replace("_", " ")} Procurement',
                'Phone': f'021-{i:04d}-{i*111:04d}',
                'Email': f'{destination.lower()}@hospital.co.id',
                'Status': 'Active'
            })

        df = pd.DataFrame(data)
        df.to_csv(filename, index=False)

        print(f"✓ Generated: {filename}")
        print(f"  Rows: {len(df)}")
        print(f"  Columns: {len(df.columns)}")

        return df

    def generate_transportation_cost_csv(self):
        """Generate transportation_cost.csv"""

        filename = f'{self.data_dir}/transportation_cost.csv'

        # Create DataFrame with all routes
        data = []
        for warehouse in self.warehouses:
            for destination in self.destinations:
                cost = self.costs[(warehouse, destination)]

                # Calculate distance (simplified: cost * 5 km)
                distance_km = cost * 5

                # Estimate time (simplified: distance / 30 km/h avg speed)
                travel_time_hours = distance_km / 30

                data.append({
                    'Route_ID': f'{warehouse[:3].upper()}_{destination[:3].upper()}',
                    'From_Warehouse': warehouse,
                    'To_Destination': destination.replace('_', ' '),
                    'Distance_KM': round(distance_km, 1),
                    'Travel_Time_Hours': round(travel_time_hours, 2),
                    'Cost_Per_Unit_Rp_Thousands': cost,
                    'Fuel_Cost_Rp_Thousands': round(cost * 0.6, 2),
                    'Driver_Cost_Rp_Thousands': round(cost * 0.25, 2),
                    'Maintenance_Cost_Rp_Thousands': round(cost * 0.15, 2),
                    'Route_Condition': self._get_route_condition(cost),
                    'Traffic_Level': self._get_traffic_level(warehouse, destination),
                    'Preferred_Route': 'Yes' if cost <= 10 else 'No',
                    'Last_Updated': '2025-01-01'
                })

        df = pd.DataFrame(data)
        df.to_csv(filename, index=False)

        print(f"✓ Generated: {filename}")
        print(f"  Rows: {len(df)}")
        print(f"  Columns: {len(df.columns)}")

        return df

    def _get_route_condition(self, cost):
        """Determine route condition based on cost"""
        if cost <= 7:
            return 'Excellent'
        elif cost <= 15:
            return 'Good'
        elif cost <= 20:
            return 'Fair'
        else:
            return 'Poor'

    def _get_traffic_level(self, warehouse, destination):
        """Determine traffic level based on locations"""
        high_traffic = ['Jakarta', 'Tangerang']

        if warehouse in high_traffic or any(loc in destination for loc in high_traffic):
            return 'High'
        elif 'Bekasi' in warehouse or 'Bekasi' in destination:
            return 'Medium'
        else:
            return 'Low'

    def generate_input_data_xlsx(self):
        """Generate comprehensive input_data.xlsx with multiple sheets"""

        filename = f'{self.data_dir}/input_data.xlsx'

        # Create Excel workbook
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="366092")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Overview
        ws_overview = wb.create_sheet("Overview", 0)
        self._create_overview_sheet(ws_overview, title_font, header_fill, header_font)

        # Sheet 2: Warehouse Capacity
        ws_warehouse = wb.create_sheet("Warehouse_Capacity", 1)
        self._create_warehouse_sheet(ws_warehouse, header_fill, header_font, border)

        # Sheet 3: Destination Demand
        ws_destination = wb.create_sheet("Destination_Demand", 2)
        self._create_destination_sheet(ws_destination, header_fill, header_font, border)

        # Sheet 4: Transportation Costs
        ws_costs = wb.create_sheet("Transportation_Costs", 3)
        self._create_costs_sheet(ws_costs, header_fill, header_font, border)

        # Sheet 5: Cost Matrix
        ws_matrix = wb.create_sheet("Cost_Matrix", 4)
        self._create_cost_matrix_sheet(ws_matrix, header_fill, header_font, border)

        # Sheet 6: Summary Statistics
        ws_stats = wb.create_sheet("Summary_Statistics", 5)
        self._create_statistics_sheet(ws_stats, header_fill, header_font, border)

        # Save workbook
        wb.save(filename)

        print(f"✓ Generated: {filename}")
        print(f"  Sheets: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")

        return filename

    def _create_overview_sheet(self, ws, title_font, header_fill, header_font):
        """Create overview sheet"""

        # Title
        ws['A1'] = "Transportation Problem - PT. MediCare Indonesia"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        # Problem description
        ws['A3'] = "Problem Description:"
        ws['A3'].font = Font(bold=True, size=12)

        description = [
            "This dataset contains information for optimizing pharmaceutical product distribution",
            "from multiple warehouses to various healthcare facilities in the Jabodetabek region.",
            "",
            "Objective: Minimize total transportation cost while satisfying all demand",
            "",
            "Data Date: January 2026",
            "Status: Active",
            "Problem Type: Transportation Problem (Linear Programming)"
        ]

        for i, line in enumerate(description, 4):
            ws[f'A{i}'] = line

        # Key metrics
        row = 13
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        metrics = [
            ['Metric', 'Value', 'Unit'],
            ['Total Warehouses', len(self.warehouses), 'facilities'],
            ['Total Destinations', len(self.destinations), 'facilities'],
            ['Total Supply Capacity', sum(self.supply.values()), 'units'],
            ['Total Demand', sum(self.demand.values()), 'units'],
            ['Supply-Demand Balance', sum(self.supply.values()) - sum(self.demand.values()), 'units'],
            ['Number of Routes', len(self.costs), 'routes'],
            ['Min Cost per Unit', min(self.costs.values()), 'Rp thousands'],
            ['Max Cost per Unit', max(self.costs.values()), 'Rp thousands'],
            ['Avg Cost per Unit', round(sum(self.costs.values()) / len(self.costs), 2), 'Rp thousands']
        ]

        for i, metric in enumerate(metrics):
            for j, value in enumerate(metric, 1):
                cell = ws.cell(row + i, j, value)
                if i == 0:
                    cell.fill = header_fill
                    cell.font = header_font

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30

    def _create_warehouse_sheet(self, ws, header_fill, header_font, border):
        """Create warehouse capacity sheet"""

        # Headers
        headers = ['Warehouse_ID', 'Warehouse_Name', 'Location', 'Capacity_Units',
                   'Status', 'Manager', 'Contact']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        row = 2
        for warehouse in self.warehouses:
            ws.cell(row, 1, f'WH_{warehouse[:3].upper()}')
            ws.cell(row, 2, warehouse)
            ws.cell(row, 3, warehouse)
            ws.cell(row, 4, self.supply[warehouse])
            ws.cell(row, 5, 'Active')
            ws.cell(row, 6, f'{warehouse} Manager')
            ws.cell(row, 7, f'{warehouse.lower()}@medicare.co.id')
            row += 1

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_destination_sheet(self, ws, header_fill, header_font, border):
        """Create destination demand sheet"""

        # Headers
        headers = ['Destination_ID', 'Destination_Name', 'Facility_Type',
                   'Monthly_Demand_Units', 'Priority_Level', 'Contact_Person']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        row = 2
        for i, destination in enumerate(self.destinations, 1):
            facility_type = 'Hospital' if 'RS' in destination else 'Pharmacy'
            priority = 'High' if self.demand[destination] >= 250 else 'Medium'

            ws.cell(row, 1, f'DEST_{i:03d}')
            ws.cell(row, 2, destination.replace('_', ' '))
            ws.cell(row, 3, facility_type)
            ws.cell(row, 4, self.demand[destination])
            ws.cell(row, 5, priority)
            ws.cell(row, 6, f'{destination.replace("_", " ")} Procurement')
            row += 1

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _create_costs_sheet(self, ws, header_fill, header_font, border):
        """Create transportation costs sheet"""

        # Headers
        headers = ['Route_ID', 'From_Warehouse', 'To_Destination', 'Distance_KM',
                   'Cost_Per_Unit_Rp_Thousands', 'Route_Condition', 'Traffic_Level']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        row = 2
        for warehouse in self.warehouses:
            for destination in self.destinations:
                cost = self.costs[(warehouse, destination)]
                distance = cost * 5
                condition = self._get_route_condition(cost)
                traffic = self._get_traffic_level(warehouse, destination)

                ws.cell(row, 1, f'{warehouse[:3].upper()}_{destination[:3].upper()}')
                ws.cell(row, 2, warehouse)
                ws.cell(row, 3, destination.replace('_', ' '))
                ws.cell(row, 4, round(distance, 1))
                ws.cell(row, 5, cost)
                ws.cell(row, 6, condition)
                ws.cell(row, 7, traffic)
                row += 1

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    def _create_cost_matrix_sheet(self, ws, header_fill, header_font, border):
        """Create cost matrix sheet"""

        # Title
        ws['A1'] = "Cost Matrix (Rp thousands per unit)"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')

        # Headers
        ws.cell(3, 1, "From \\ To").fill = header_fill
        ws.cell(3, 1).font = header_font

        for j, dest in enumerate(self.destinations, 2):
            cell = ws.cell(3, j, dest.replace('_', '\n'))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Data
        row = 4
        for warehouse in self.warehouses:
            cell = ws.cell(row, 1, warehouse)
            cell.fill = header_fill
            cell.font = header_font

            for j, dest in enumerate(self.destinations, 2):
                cost = self.costs[(warehouse, dest)]
                cell = ws.cell(row, j, cost)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

                # Color code by cost
                if cost <= 7:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cost <= 15:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(self.destinations) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Row height for wrapped headers
        ws.row_dimensions[3].height = 30

    def _create_statistics_sheet(self, ws, header_fill, header_font, border):
        """Create summary statistics sheet"""

        # Title
        ws['A1'] = "Summary Statistics"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')

        # Supply statistics
        row = 3
        ws.cell(row, 1, "Supply Statistics").font = Font(bold=True, size=12)
        row += 1

        supply_values = list(self.supply.values())
        stats = [
            ['Metric', 'Value', 'Unit'],
            ['Total Supply', sum(supply_values), 'units'],
            ['Average Capacity', round(np.mean(supply_values), 2), 'units'],
            ['Std Deviation', round(np.std(supply_values), 2), 'units'],
            ['Min Capacity', min(supply_values), 'units'],
            ['Max Capacity', max(supply_values), 'units']
        ]

        for i, stat in enumerate(stats):
            for j, value in enumerate(stat, 1):
                cell = ws.cell(row + i, j, value)
                if i == 0:
                    cell.fill = header_fill
                    cell.font = header_font

        # Demand statistics
        row += len(stats) + 2
        ws.cell(row, 1, "Demand Statistics").font = Font(bold=True, size=12)
        row += 1

        demand_values = list(self.demand.values())
        stats = [
            ['Metric', 'Value', 'Unit'],
            ['Total Demand', sum(demand_values), 'units'],
            ['Average Demand', round(np.mean(demand_values), 2), 'units'],
            ['Std Deviation', round(np.std(demand_values), 2), 'units'],
            ['Min Demand', min(demand_values), 'units'],
            ['Max Demand', max(demand_values), 'units']
        ]

        for i, stat in enumerate(stats):
            for j, value in enumerate(stat, 1):
                cell = ws.cell(row + i, j, value)
                if i == 0:
                    cell.fill = header_fill
                    cell.font = header_font

        # Cost statistics
        row += len(stats) + 2
        ws.cell(row, 1, "Cost Statistics").font = Font(bold=True, size=12)
        row += 1

        cost_values = list(self.costs.values())
        stats = [
            ['Metric', 'Value', 'Unit'],
            ['Average Cost', round(np.mean(cost_values), 2), 'Rp thousands'],
            ['Std Deviation', round(np.std(cost_values), 2), 'Rp thousands'],
            ['Min Cost', min(cost_values), 'Rp thousands'],
            ['Max Cost', max(cost_values), 'Rp thousands'],
            ['Total Routes', len(cost_values), 'routes']
        ]

        for i, stat in enumerate(stats):
            for j, value in enumerate(stat, 1):
                cell = ws.cell(row + i, j, value)
                if i == 0:
                    cell.fill = header_fill
                    cell.font = header_font

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

    def generate_all_files(self):
        """Generate all data files"""

        print("="*70)
        print("GENERATING ALL DATA FILES")
        print("="*70)
        print()

        # Generate CSV files
        self.generate_warehouse_capacity_csv()
        print()

        self.generate_destination_demand_csv()
        print()

        self.generate_transportation_cost_csv()
        print()

        # Generate Excel file
        self.generate_input_data_xlsx()
        print()

        print("="*70)
        print("ALL DATA FILES GENERATED SUCCESSFULLY!")
        print("="*70)
        print()
        print("Generated files:")
        print("  1. warehouse_capacity.csv")
        print("  2. destination_demand.csv")
        print("  3. transportation_cost.csv")
        print("  4. input_data.xlsx (6 sheets)")
        print()
        print(f"Location: {self.data_dir}/")
        print("="*70)


# Main execution
if __name__ == "__main__":
    # Create generator
    generator = DataGenerator()

    # Generate all files
    generator.generate_all_files()

    print("\n✓ Data generation complete!")
    print("\nNext steps:")
    print("  1. Check the generated files in '../data/' directory")
    print("  2. Review the data for accuracy")
    print("  3. Proceed to run the notebooks")
    print("  4. Use these files as input for your analysis")