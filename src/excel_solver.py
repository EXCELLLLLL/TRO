"""
Excel Solver Helper Module
PT. MediCare Indonesia - Transportation Problem

This module provides utilities to:
1. Generate Excel files formatted for Solver
2. Read Solver results from Excel
3. Validate Excel Solver setup
4. Export results for comparison

Note: This module prepares Excel files but actual solving
must be done manually in Excel using the Solver add-in.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from model_formulation import TransportationData
import os


class ExcelSolverHelper:
    """Helper class for Excel Solver setup and result processing"""

    def __init__(self, data: TransportationData):
        """
        Initialize with transportation data

        Args:
            data: TransportationData instance
        """
        self.data = data
        self.workbook = None
        self.worksheet = None

    def create_solver_workbook(self, filename='transportation_solver_template.xlsx'):
        """
        Create Excel workbook formatted for Solver

        Args:
            filename: Output filename
        """
        print(f"Creating Excel Solver template: {filename}")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Solver Model"

        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        data_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        result_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        result_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = "Transportation Problem - PT. MediCare Indonesia"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')

        # Section 1: Cost Matrix
        row = 3
        ws.cell(row, 1, "COST MATRIX (Rp thousands per unit)").font = Font(bold=True, size=12)
        row += 1

        # Headers
        ws.cell(row, 1, "From \\ To").fill = header_fill
        ws.cell(row, 1).font = header_font
        for j, dest in enumerate(self.data.destinations, 2):
            ws.cell(row, j, dest.replace('_', ' ')).fill = header_fill
            ws.cell(row, j).font = header_font
            ws.cell(row, j).alignment = Alignment(horizontal='center')

        # Cost data
        row += 1
        cost_start_row = row
        for i, warehouse in enumerate(self.data.warehouses):
            ws.cell(row, 1, warehouse).fill = header_fill
            ws.cell(row, 1).font = header_font

            for j, dest in enumerate(self.data.destinations, 2):
                cost = self.data.costs[(warehouse, dest)]
                cell = ws.cell(row, j, cost)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            row += 1
        cost_end_row = row - 1

        # Section 2: Supply and Demand
        row += 1
        supply_row_start = row

        # Supply
        ws.cell(row, 1, "SUPPLY (Capacity)").font = Font(bold=True, size=12)
        ws.cell(row, 4, "DEMAND (Requirements)").font = Font(bold=True, size=12)
        row += 1

        ws.cell(row, 1, "Warehouse").fill = header_fill
        ws.cell(row, 1).font = header_font
        ws.cell(row, 2, "Capacity").fill = header_fill
        ws.cell(row, 2).font = header_font

        ws.cell(row, 4, "Destination").fill = header_fill
        ws.cell(row, 4).font = header_font
        ws.cell(row, 5, "Demand").fill = header_fill
        ws.cell(row, 5).font = header_font

        row += 1
        supply_data_start = row

        for warehouse in self.data.warehouses:
            ws.cell(row, 1, warehouse)
            ws.cell(row, 2, self.data.supply[warehouse]).alignment = Alignment(horizontal='center')
            row += 1
        supply_data_end = row - 1

        # Demand (align with supply)
        row = supply_data_start
        demand_data_start = row
        for dest in self.data.destinations:
            ws.cell(row, 4, dest.replace('_', ' '))
            ws.cell(row, 5, self.data.demand[dest]).alignment = Alignment(horizontal='center')
            row += 1
        demand_data_end = row - 1

        # Section 3: Decision Variables (Allocation Matrix)
        row += 2
        ws.cell(row, 1, "ALLOCATION MATRIX (Decision Variables)").font = Font(bold=True, size=12)
        ws.cell(row, 8, "← Solver will change these cells").font = Font(italic=True, color="FF0000")
        row += 1

        # Headers
        ws.cell(row, 1, "From \\ To").fill = header_fill
        ws.cell(row, 1).font = header_font
        for j, dest in enumerate(self.data.destinations, 2):
            ws.cell(row, j, dest.replace('_', ' ')).fill = header_fill
            ws.cell(row, j).font = header_font
            ws.cell(row, j).alignment = Alignment(horizontal='center')
        ws.cell(row, len(self.data.destinations) + 2, "Total Shipped").fill = header_fill
        ws.cell(row, len(self.data.destinations) + 2).font = header_font

        row += 1
        allocation_start_row = row

        # Allocation cells (initially 0)
        for i, warehouse in enumerate(self.data.warehouses):
            ws.cell(row, 1, warehouse).fill = header_fill
            ws.cell(row, 1).font = header_font

            for j, dest in enumerate(self.data.destinations, 2):
                cell = ws.cell(row, j, 0)
                cell.fill = data_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            # Total shipped (sum formula)
            col = len(self.data.destinations) + 2
            start_col = get_column_letter(2)
            end_col = get_column_letter(len(self.data.destinations) + 1)
            ws.cell(row, col, f"=SUM({start_col}{row}:{end_col}{row})")
            ws.cell(row, col).fill = result_fill
            ws.cell(row, col).alignment = Alignment(horizontal='center')

            row += 1
        allocation_end_row = row - 1

        # Total received row
        ws.cell(row, 1, "Total Received").fill = header_fill
        ws.cell(row, 1).font = header_font

        for j, dest in enumerate(self.data.destinations, 2):
            col_letter = get_column_letter(j)
            ws.cell(row, j, f"=SUM({col_letter}{allocation_start_row}:{col_letter}{allocation_end_row})")
            ws.cell(row, j).fill = result_fill
            ws.cell(row, j).alignment = Alignment(horizontal='center')

        total_row = row

        # Section 4: Objective Function
        row += 2
        ws.cell(row, 1, "OBJECTIVE FUNCTION").font = Font(bold=True, size=12)
        row += 1

        ws.cell(row, 1, "Total Transportation Cost:").font = Font(bold=True)

        # SUMPRODUCT formula
        cost_range = f"B{cost_start_row}:{get_column_letter(len(self.data.destinations)+1)}{cost_end_row}"
        alloc_range = f"B{allocation_start_row}:{get_column_letter(len(self.data.destinations)+1)}{allocation_end_row}"

        ws.cell(row, 2, f"=SUMPRODUCT({cost_range},{alloc_range})")
        ws.cell(row, 2).fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        ws.cell(row, 2).font = Font(bold=True, size=14)
        ws.cell(row, 2).number_format = '#,##0'

        ws.cell(row, 3, "Rp thousands").font = Font(italic=True)

        objective_cell = f"B{row}"

        # Section 5: Solver Instructions
        row += 3
        ws.cell(row, 1, "SOLVER SETUP INSTRUCTIONS:").font = Font(bold=True, size=12, color="FF0000")
        row += 1

        instructions = [
            "1. Go to Data → Solver (if not visible, enable Solver Add-in)",
            f"2. Set Objective: {objective_cell} (Min)",
            f"3. By Changing Variable Cells: B{allocation_start_row}:{get_column_letter(len(self.data.destinations)+1)}{allocation_end_row}",
            "4. Add Constraints:",
            f"   a) Total Shipped ≤ Supply: {get_column_letter(len(self.data.destinations)+2)}{allocation_start_row}:{get_column_letter(len(self.data.destinations)+2)}{allocation_end_row} <= B{supply_data_start}:B{supply_data_end}",
            f"   b) Total Received = Demand: B{total_row}:{get_column_letter(len(self.data.destinations)+1)}{total_row} = E{demand_data_start}:E{demand_data_end}",
            f"   c) Non-negative: B{allocation_start_row}:{get_column_letter(len(self.data.destinations)+1)}{allocation_end_row} >= 0",
            "5. Select Solving Method: Simplex LP",
            "6. Click Options → Check 'Make Unconstrained Variables Non-Negative'",
            "7. Click Solve",
            "8. When done, select 'Keep Solver Solution' and generate all reports"
        ]

        for instruction in instructions:
            ws.cell(row, 1, instruction).font = Font(size=10)
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for i in range(2, len(self.data.destinations) + 5):
            ws.column_dimensions[get_column_letter(i)].width = 15

        # Save workbook
        wb.save(filename)
        print(f"✓ Excel template created: {filename}")
        print(f"  Open in Excel and follow the instructions to run Solver")

        return filename

    def read_solver_results(self, filename):
        """
        Read results after Solver has been run in Excel

        Args:
            filename: Excel file with Solver results

        Returns:
            dict: Results including allocation, cost, and statistics
        """
        print(f"Reading Solver results from: {filename}")

        if not os.path.exists(filename):
            raise FileNotFoundError(f"File not found: {filename}")

        # Load workbook
        wb = load_workbook(filename, data_only=True)
        ws = wb.active

        # Find allocation matrix
        # This is a simplified version - adjust based on actual layout
        allocation = {}

        # Read allocation data (assuming standard layout)
        # You would need to adjust these based on actual Excel layout
        allocation_start_row = 15  # Adjust based on template

        for i, warehouse in enumerate(self.data.warehouses):
            for j, dest in enumerate(self.data.destinations):
                row = allocation_start_row + i
                col = 2 + j  # Assuming allocation starts at column B
                value = ws.cell(row, col).value

                if value and value > 0:
                    allocation[(warehouse, dest)] = float(value)

        # Read objective value
        # Adjust cell reference based on template
        objective_value = ws.cell(20, 2).value  # Example cell

        results = {
            'allocation': allocation,
            'total_cost': objective_value,
            'num_routes': len(allocation),
            'status': 'Read from Excel'
        }

        print(f"✓ Results loaded")
        print(f"  Total Cost: Rp {objective_value:,.0f},000")
        print(f"  Active Routes: {len(allocation)}")

        return results

    def create_comparison_workbook(self, vam_results, excel_results, python_results,
                                   filename='method_comparison.xlsx'):
        """
        Create comparison workbook for all three methods

        Args:
            vam_results: Results from VAM
            excel_results: Results from Excel Solver
            python_results: Results from Python
            filename: Output filename
        """
        print(f"Creating comparison workbook: {filename}")

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Summary comparison
            summary_data = {
                'Method': ['VAM (Manual)', 'Excel Solver', 'Python (PuLP)'],
                'Total Cost (Rp ribu)': [
                    vam_results.get('total_cost', 0),
                    excel_results.get('total_cost', 0),
                    python_results.get('total_cost', 0)
                ],
                'Status': [
                    vam_results.get('status', 'Unknown'),
                    excel_results.get('status', 'Unknown'),
                    python_results.get('status', 'Unknown')
                ],
                'Active Routes': [
                    vam_results.get('num_routes', 0),
                    excel_results.get('num_routes', 0),
                    python_results.get('num_routes', 0)
                ],
                'Solve Time': [
                    vam_results.get('solve_time', '~15 min'),
                    excel_results.get('solve_time', '<1 sec'),
                    python_results.get('solve_time', '<1 sec')
                ]
            }

            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)

            # Detailed allocation for each method
            for method_name, results in [
                ('VAM', vam_results),
                ('Excel', excel_results),
                ('Python', python_results)
            ]:
                allocation = results.get('allocation', {})

                # Create allocation matrix
                alloc_data = []
                for w in self.data.warehouses:
                    row = {'Warehouse': w}
                    for d in self.data.destinations:
                        value = allocation.get((w, d), 0)
                        row[d] = value if value > 0 else 0
                    alloc_data.append(row)

                df_alloc = pd.DataFrame(alloc_data)
                df_alloc.to_excel(writer, sheet_name=f'{method_name}_Allocation', index=False)

        print(f"✓ Comparison workbook created: {filename}")

    def validate_solver_setup(self, filename):
        """
        Validate that Excel file is correctly set up for Solver

        Args:
            filename: Excel file to validate

        Returns:
            tuple: (is_valid, list of issues)
        """
        print(f"Validating Solver setup: {filename}")

        issues = []

        try:
            wb = load_workbook(filename, data_only=False)
            ws = wb.active

            # Check if objective function exists
            # This is simplified - actual validation would be more comprehensive
            objective_cell = ws['B20']
            if not objective_cell.value or not str(objective_cell.value).startswith('='):
                issues.append("Objective function not found or invalid")

            # Check for allocation matrix
            allocation_start = ws['B15']
            if allocation_start.value is None:
                issues.append("Allocation matrix appears empty")

            # More validation checks could be added here

        except Exception as e:
            issues.append(f"Error reading file: {str(e)}")

        is_valid = len(issues) == 0

        if is_valid:
            print("✓ Solver setup is valid")
        else:
            print("✗ Issues found:")
            for issue in issues:
                print(f"  - {issue}")

        return is_valid, issues


# Example usage
if __name__ == "__main__":
    print("="*70)
    print("EXCEL SOLVER HELPER")
    print("="*70)

    # Load data
    data = TransportationData()

    # Create helper
    helper = ExcelSolverHelper(data)

    # Generate Excel template
    template_file = '../data/transportation_solver_template.xlsx'
    helper.create_solver_workbook(template_file)

    print("\n" + "="*70)
    print("NEXT STEPS:")
    print("="*70)
    print(f"1. Open {template_file} in Excel")
    print("2. Follow the instructions in the worksheet")
    print("3. Run Solver")
    print("4. Save the file with results")
    print("5. Use read_solver_results() to analyze")
    print("="*70)

    print("\n✓ Excel Solver helper ready!")